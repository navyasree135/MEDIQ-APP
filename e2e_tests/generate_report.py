"""
MediQ E2E Test Report Generator
================================
Generates an Excel (.xlsx) report matching the reference format:
  Sheet 1 — Summary (Test Suite, Total Tests, Passed, Failed, Pass Rate %, Duration, Start/End Time)
  Sheet 2 — Detailed Results (Test ID, Test Name, Module, Status, Duration, Error Message)
"""
import subprocess
import sys
import json
import os
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def run_tests_and_collect():
    """Run pytest with JSON report and return results."""
    report_json = Path(__file__).parent / "test_results.json"

    cmd = [
        sys.executable, "-m", "pytest",
        str(Path(__file__).parent / "test_mediq_e2e.py"),
        f"--json-report-file={report_json}",
        "--json-report",
        "-v",
        "--tb=short",
        "-q",
    ]

    start_time = datetime.utcnow()
    result = subprocess.run(cmd, capture_output=True, text=True, cwd=str(Path(__file__).parent))
    end_time = datetime.utcnow()

    stdout = result.stdout
    stderr = result.stderr

    # Parse JSON report if available
    if report_json.exists():
        with open(report_json, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data, start_time, end_time, stdout, stderr

    # Fallback: parse stdout
    return None, start_time, end_time, stdout, stderr


def parse_pytest_output(stdout):
    """Parse pytest verbose output to extract test results."""
    results = []
    for line in stdout.splitlines():
        # Match lines like: test_mediq_e2e.py::TestClass::test_name PASSED
        match = re.match(r".*?::(Test\w+)::(test_\w+)\s+(PASSED|FAILED|ERROR|SKIPPED)", line)
        if match:
            module = match.group(1)
            name = match.group(2)
            status = match.group(3)
            results.append({
                "nodeid": f"{module}::{name}",
                "module": module,
                "test_name": name,
                "outcome": status.lower(),
                "duration": 0,
                "message": "",
            })
    return results


def generate_xlsx(json_data, start_time, end_time, stdout, stderr):
    """Generate the Excel report."""
    wb = Workbook()

    # ── Colors & Styles ──
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    pass_fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
    fail_fill = PatternFill(start_color="FDECEA", end_color="FDECEA", fill_type="solid")
    skip_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
    pass_font = Font(color="1B7A2B", bold=True, name="Calibri")
    fail_font = Font(color="CC0000", bold=True, name="Calibri")
    skip_font = Font(color="B8860B", bold=True, name="Calibri")
    normal_font = Font(name="Calibri", size=11)
    bold_font = Font(name="Calibri", size=11, bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    center = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ── Parse test data ──
    test_items = []
    total_duration = 0

    from datetime import timedelta
    current_time = start_time

    if json_data and "tests" in json_data:
        for t in json_data["tests"]:
            nodeid = t.get("nodeid", "")
            parts = nodeid.split("::")
            module = parts[1] if len(parts) > 1 else "Unknown"
            test_name = parts[-1] if parts else "Unknown"
            outcome = t.get("outcome", "unknown")
            duration = t.get("duration", 0)
            total_duration += duration

            # Get error message if failed
            msg = ""
            if outcome == "failed":
                call = t.get("call", {})
                crash = call.get("crash", {})
                msg = crash.get("message", "")
                if not msg:
                    longrepr = call.get("longrepr", "")
                    if longrepr:
                        msg = str(longrepr)[:200]

            timestamp_str = current_time.strftime("%Y-%m-%d %H:%M:%S")
            current_time += timedelta(seconds=duration)

            test_items.append({
                "module": module,
                "test_name": test_name,
                "outcome": outcome,
                "duration": round(duration, 3),
                "timestamp": timestamp_str,
                "message": msg,
            })
    else:
        # Fallback parsing from stdout
        test_items = parse_pytest_output(stdout)
        if not test_items:
            # If we still have nothing, parse the summary line
            summary_match = re.search(r"(\d+) passed", stdout)
            failed_match = re.search(r"(\d+) failed", stdout)
            skipped_match = re.search(r"(\d+) skipped", stdout)

            passed_count = int(summary_match.group(1)) if summary_match else 0
            failed_count = int(failed_match.group(1)) if failed_match else 0
            skipped_count = int(skipped_match.group(1)) if skipped_match else 0

            for i in range(passed_count):
                test_items.append({
                    "module": "TestSuite",
                    "test_name": f"test_{i+1:03d}",
                    "outcome": "passed",
                    "duration": 0,
                    "timestamp": current_time.strftime("%Y-%m-%d %H:%M:%S"),
                    "message": "",
                })
            for i in range(failed_count):
                test_items.append({
                    "module": "TestSuite",
                    "test_name": f"test_failed_{i+1:03d}",
                    "outcome": "failed",
                    "duration": 0,
                    "message": "See test logs",
                })
            for i in range(skipped_count):
                test_items.append({
                    "module": "TestSuite",
                    "test_name": f"test_skipped_{i+1:03d}",
                    "outcome": "skipped",
                    "duration": 0,
                    "message": "Skipped",
                })

    total = len(test_items)
    passed = sum(1 for t in test_items if t["outcome"] == "passed")
    failed = sum(1 for t in test_items if t["outcome"] == "failed")
    skipped = sum(1 for t in test_items if t["outcome"] == "skipped")
    errors = sum(1 for t in test_items if t["outcome"] == "error")
    pass_rate = round((passed / total) * 100, 2) if total > 0 else 0
    duration_sec = round((end_time - start_time).total_seconds(), 2)

    # ══════════════════════════════════════════
    # SHEET 1 — Summary
    # ══════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Summary"

    summary_headers = [
        "Test Suite", "Total Tests", "Passed", "Failed", "Skipped",
        "Pass Rate %", "Duration (sec)", "Start Time", "End Time"
    ]
    for col_idx, header in enumerate(summary_headers, 1):
        cell = ws1.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border

    summary_values = [
        "MediQ Healthcare App — Full E2E Workflow",
        total, passed, failed, skipped,
        pass_rate, duration_sec,
        start_time.isoformat() + "Z",
        end_time.isoformat() + "Z",
    ]
    for col_idx, val in enumerate(summary_values, 1):
        cell = ws1.cell(row=2, column=col_idx, value=val)
        cell.font = normal_font
        cell.alignment = center
        cell.border = thin_border

    # Auto-width
    for col_idx in range(1, len(summary_headers) + 1):
        ws1.column_dimensions[get_column_letter(col_idx)].width = max(18, len(str(summary_headers[col_idx - 1])) + 6)

    # ══════════════════════════════════════════
    # SHEET 2 — Detailed Results
    # ══════════════════════════════════════════
    ws2 = wb.create_sheet(title="Test Results")

    detail_headers = [
        "S.No", "Test ID", "Test Name", "Module", "Status",
        "Duration (sec)", "Timestamp", "Error / Message"
    ]
    for col_idx, header in enumerate(detail_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border

    for row_idx, item in enumerate(test_items, 2):
        # S.No
        ws2.cell(row=row_idx, column=1, value=row_idx - 1).font = normal_font
        ws2.cell(row=row_idx, column=1).alignment = center
        ws2.cell(row=row_idx, column=1).border = thin_border

        # Test ID
        test_id_match = re.search(r"test_(\d+)", item["test_name"])
        test_id = f"TC-{test_id_match.group(1)}" if test_id_match else f"TC-{row_idx - 1:03d}"
        ws2.cell(row=row_idx, column=2, value=test_id).font = bold_font
        ws2.cell(row=row_idx, column=2).alignment = center
        ws2.cell(row=row_idx, column=2).border = thin_border

        # Test Name — use docstring-style clean name
        clean_name = item["test_name"].replace("test_", "").replace("_", " ").title()
        ws2.cell(row=row_idx, column=3, value=clean_name).font = normal_font
        ws2.cell(row=row_idx, column=3).alignment = left_align
        ws2.cell(row=row_idx, column=3).border = thin_border

        # Module
        module = item["module"].replace("Test", "")
        ws2.cell(row=row_idx, column=4, value=module).font = normal_font
        ws2.cell(row=row_idx, column=4).alignment = center
        ws2.cell(row=row_idx, column=4).border = thin_border

        # Status
        outcome = item["outcome"].upper()
        status_cell = ws2.cell(row=row_idx, column=5, value=outcome)
        status_cell.alignment = center
        status_cell.border = thin_border
        if outcome == "PASSED":
            status_cell.fill = pass_fill
            status_cell.font = pass_font
        elif outcome == "FAILED" or outcome == "ERROR":
            status_cell.fill = fail_fill
            status_cell.font = fail_font
        elif outcome == "SKIPPED":
            status_cell.fill = skip_fill
            status_cell.font = skip_font

        # Duration
        ws2.cell(row=row_idx, column=6, value=item["duration"]).font = normal_font
        ws2.cell(row=row_idx, column=6).alignment = center
        ws2.cell(row=row_idx, column=6).border = thin_border

        # Timestamp
        ws2.cell(row=row_idx, column=7, value=item.get("timestamp", "")).font = normal_font
        ws2.cell(row=row_idx, column=7).alignment = center
        ws2.cell(row=row_idx, column=7).border = thin_border

        # Error Message
        msg = item.get("message", "")[:300]
        ws2.cell(row=row_idx, column=8, value=msg).font = normal_font
        ws2.cell(row=row_idx, column=8).alignment = left_align
        ws2.cell(row=row_idx, column=8).border = thin_border

    # Auto-width for detail sheet
    col_widths = [8, 10, 55, 25, 12, 14, 20, 60]
    for col_idx, w in enumerate(col_widths, 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = w

    # ── Save ──
    timestamp = datetime.now().strftime("%Y-%m-%dT%H-%M-%S")
    filename = f"E2E_Test_Report_MediQ_{timestamp}.xlsx"
    output_path = Path(__file__).parent.parent / filename
    wb.save(str(output_path))
    print(f"\n{'='*60}")
    print(f"  [OK] E2E TEST REPORT GENERATED")
    print(f"  File: {output_path}")
    print(f"  Total: {total} | Passed: {passed} | Failed: {failed} | Skipped: {skipped}")
    print(f"  Pass Rate: {pass_rate}%")
    print(f"  Duration: {duration_sec}s")
    print(f"{'='*60}\n")
    return str(output_path)


if __name__ == "__main__":
    # Try to use json-report plugin, if not available fall back
    try:
        import pytest_jsonreport
    except ImportError:
        print("Installing pytest-json-report for structured results...")
        subprocess.run([sys.executable, "-m", "pip", "install", "pytest-json-report", "-q"], check=False)

    json_data, start_time, end_time, stdout, stderr = run_tests_and_collect()
    print(stdout)
    if stderr:
        print(stderr)
    generate_xlsx(json_data, start_time, end_time, stdout, stderr)
