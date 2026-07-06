"""
MediQ Appium Test Report Generator
==================================
Runs the Appium E2E test suite and outputs the results in a styled Excel (.xlsx) file.
"""
import subprocess
import sys
import json
import os
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def run_tests_and_collect():
    """Run pytest with JSON report option."""
    report_json = Path(__file__).parent / "appium_test_results.json"
    
    cmd = [
        sys.executable, "-m", "pytest",
        str(Path(__file__).parent / "test_mediq_appium.py"),
        f"--json-report-file={report_json}",
        "--json-report",
        "-v",
        "--tb=short",
        "-q"
    ]
    
    start_time = datetime.utcnow()
    result = subprocess.run(cmd, capture_output=True, text=True, cwd=str(Path(__file__).parent))
    end_time = datetime.utcnow()
    
    stdout = result.stdout
    stderr = result.stderr
    
    if report_json.exists():
        with open(report_json, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data, start_time, end_time, stdout, stderr
        
    return None, start_time, end_time, stdout, stderr

def parse_pytest_output(stdout):
    results = []
    for line in stdout.splitlines():
        match = re.match(r".*?::(Test\w+)::(test_\w+)\s+(PASSED|FAILED|ERROR|SKIPPED)", line)
        if match:
            module = match.group(1)
            name = match.group(2)
            status = match.group(3)
            results.append({
                "nodeid": f"{module}::{name}",
                "module": module,
                "test_name": name,
                "outcome": status.lower(),
                "duration": 0,
                "message": ""
            })
    return results

def generate_xlsx(json_data, start_time, end_time, stdout, stderr):
    wb = Workbook()
    
    # styles
    header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    pass_fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
    fail_fill = PatternFill(start_color="FDECEA", end_color="FDECEA", fill_type="solid")
    skip_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
    
    pass_font = Font(color="1B7A2B", bold=True, name="Calibri")
    fail_font = Font(color="CC0000", bold=True, name="Calibri")
    skip_font = Font(color="B8860B", bold=True, name="Calibri")
    
    normal_font = Font(name="Calibri", size=11)
    bold_font = Font(name="Calibri", size=11, bold=True)
    
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0")
    )
    
    center = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    test_items = []
    total_duration = 0
    
    from datetime import timedelta
    current_time = start_time
    
    if json_data and "tests" in json_data:
        for t in json_data["tests"]:
            nodeid = t.get("nodeid", "")
            parts = nodeid.split("::")
            module = parts[1] if len(parts) > 1 else "MobileApp"
            test_name = parts[-1] if parts else "Unknown"
            outcome = t.get("outcome", "unknown")
            duration = t.get("duration", 0)
            total_duration += duration
            
            msg = ""
            if outcome == "failed":
                call = t.get("call", {})
                crash = call.get("crash", {})
                msg = crash.get("message", "")
                if not msg:
                    longrepr = call.get("longrepr", "")
                    if longrepr:
                        msg = str(longrepr)[:200]
                        
            timestamp_str = current_time.strftime("%Y-%m-%d %H:%M:%S")
            current_time += timedelta(seconds=duration)
            
            test_items.append({
                "module": module,
                "test_name": test_name,
                "outcome": outcome,
                "duration": round(duration, 3),
                "timestamp": timestamp_str,
                "message": msg
            })
    else:
        test_items = parse_pytest_output(stdout)
        
    total = len(test_items)
    passed = sum(1 for t in test_items if t["outcome"] == "passed")
    failed = sum(1 for t in test_items if t["outcome"] == "failed")
    skipped = sum(1 for t in test_items if t["outcome"] == "skipped")
    pass_rate = round((passed / total) * 100, 2) if total > 0 else 0
    duration_sec = round((end_time - start_time).total_seconds(), 2)
    
    # Sheet 1: Summary
    ws1 = wb.active
    ws1.title = "Appium Summary"
    
    summary_headers = [
        "Mobile Test Suite", "Total Tests", "Passed", "Failed", "Skipped",
        "Pass Rate %", "Duration (sec)", "Start Time", "End Time"
    ]
    
    for col_idx, header in enumerate(summary_headers, 1):
        cell = ws1.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border
        
    summary_values = [
        "MediQ Mobile App — E2E Appium Workflow",
        total, passed, failed, skipped,
        pass_rate, duration_sec,
        start_time.isoformat() + "Z",
        end_time.isoformat() + "Z"
    ]
    
    for col_idx, val in enumerate(summary_values, 1):
        cell = ws1.cell(row=2, column=col_idx, value=val)
        cell.font = normal_font
        cell.alignment = center
        cell.border = thin_border
        
    for col_idx in range(1, len(summary_headers) + 1):
        ws1.column_dimensions[get_column_letter(col_idx)].width = max(18, len(str(summary_headers[col_idx - 1])) + 6)
        
    # Sheet 2: Detail Results
    ws2 = wb.create_sheet(title="Appium Detailed Results")
    
    detail_headers = [
        "S.No", "Test ID", "Test Name", "Module", "Status",
        "Duration (sec)", "Timestamp", "Error / Message"
    ]
    
    for col_idx, header in enumerate(detail_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border
        
    for row_idx, item in enumerate(test_items, 2):
        ws2.cell(row=row_idx, column=1, value=row_idx - 1).font = normal_font
        ws2.cell(row=row_idx, column=1).alignment = center
        ws2.cell(row=row_idx, column=1).border = thin_border
        
        test_id_match = re.search(r"test_(\d+)", item["test_name"])
        test_id = f"MOB-TC-{test_id_match.group(1)}" if test_id_match else f"MOB-TC-{row_idx - 1:03d}"
        ws2.cell(row=row_idx, column=2, value=test_id).font = bold_font
        ws2.cell(row=row_idx, column=2).alignment = center
        ws2.cell(row=row_idx, column=2).border = thin_border
        
        clean_name = item["test_name"].replace("test_", "").replace("_", " ").title()
        ws2.cell(row=row_idx, column=3, value=clean_name).font = normal_font
        ws2.cell(row=row_idx, column=3).alignment = left_align
        ws2.cell(row=row_idx, column=3).border = thin_border
        
        module = item["module"].replace("Test", "")
        ws2.cell(row=row_idx, column=4, value=module).font = normal_font
        ws2.cell(row=row_idx, column=4).alignment = center
        ws2.cell(row=row_idx, column=4).border = thin_border
        
        outcome = item["outcome"].upper()
        status_cell = ws2.cell(row=row_idx, column=5, value=outcome)
        status_cell.alignment = center
        status_cell.border = thin_border
        if outcome == "PASSED":
            status_cell.fill = pass_fill
            status_cell.font = pass_font
        elif outcome == "FAILED" or outcome == "ERROR":
            status_cell.fill = fail_fill
            status_cell.font = fail_font
        elif outcome == "SKIPPED":
            status_cell.fill = skip_fill
            status_cell.font = skip_font
            
        ws2.cell(row=row_idx, column=6, value=item["duration"]).font = normal_font
        ws2.cell(row=row_idx, column=6).alignment = center
        ws2.cell(row=row_idx, column=6).border = thin_border
        
        ws2.cell(row=row_idx, column=7, value=item.get("timestamp", "")).font = normal_font
        ws2.cell(row=row_idx, column=7).alignment = center
        ws2.cell(row=row_idx, column=7).border = thin_border
        
        ws2.cell(row=row_idx, column=8, value=item.get("message", "")[:300]).font = normal_font
        ws2.cell(row=row_idx, column=8).alignment = left_align
        ws2.cell(row=row_idx, column=8).border = thin_border
        
    col_widths = [8, 12, 55, 25, 12, 14, 20, 60]
    for col_idx, w in enumerate(col_widths, 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = w
        
    ws2.freeze_panes = "A2"
    ws1.freeze_panes = "A2"
    
    timestamp = datetime.now().strftime("%Y-%m-%dT%H-%M-%S")
    filename = f"Appium_Test_Report_MediQ_{timestamp}.xlsx"
    output_path = Path(__file__).parent.parent / filename
    wb.save(str(output_path))
    
    print(f"\n{'='*60}")
    print(f"  [OK] APPIUM E2E TEST REPORT GENERATED")
    print(f"  File: {output_path}")
    print(f"  Total: {total} | Passed: {passed} | Failed: {failed}")
    print(f"  Pass Rate: {pass_rate}%")
    print(f"{'='*60}\n")
    
    # Write summary to GitHub Step Summary if running in CI
    step_summary_path = os.environ.get("GITHUB_STEP_SUMMARY")
    if step_summary_path:
        try:
            with open(step_summary_path, "a", encoding="utf-8") as f:
                f.write("### 📱 Appium Mobile Test Execution Summary (Live)\n\n")
                f.write("| Metric | Value |\n")
                f.write("| --- | --- |\n")
                f.write(f"| **Test Suite** | MediQ Mobile App — Appium Live Suite |\n")
                f.write(f"| **Total Tests** | {total} |\n")
                f.write(f"| **Passed** | {passed} ✅ |\n")
                f.write(f"| **Failed** | {failed} ❌ |\n")
                f.write(f"| **Skipped** | {skipped} ⚠️ |\n")
                f.write(f"| **Pass Rate** | {pass_rate}% |\n")
                f.write(f"| **Duration** | {duration_sec}s ⏱️ |\n\n")
        except Exception as e:
            print(f"Error writing to GITHUB_STEP_SUMMARY: {e}")
            
    return str(output_path)

if __name__ == "__main__":
    try:
        import pytest_jsonreport
    except ImportError:
        subprocess.run([sys.executable, "-m", "pip", "install", "pytest-json-report", "-q"], check=False)
        
    json_data, start_time, end_time, stdout, stderr = run_tests_and_collect()
    print(stdout)
    if stderr:
        print(stderr)
    generate_xlsx(json_data, start_time, end_time, stdout, stderr)
